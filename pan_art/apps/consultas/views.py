import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
    ProjectedPieChart,
    PieChart
)

from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from openpyxl.chart.series import DataPoint

from django.shortcuts import render
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.blog.models import Post, Comentarios, Like, Vistas
from apps.discusion.models import (
    Opinion, 
    RespuestasAOpinion, 
    LikeAOpinion, 
    DislikeAOpinion, 
    LikeARespuestaAOpinion, 
    DislikeARespuestaAOpinion,
)
from apps.foro.models import Hilo, Comentario, LikesAHilo, DislikesAHilo
from apps.user.models import UserAccount


from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

class TablaDeUsuarios(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        usuarios = UserAccount.objects.all()
        usuario = request.user
        if usuario.is_staff:
            if usuarios.exists():
                wb = Workbook()
                ws = wb.active
                ws.title = "Usuarios"
                ws.append(["Matricula", "Nombre", "Email", "Fecha de Entrada", "Fecha de Nacimiento", "Publicaciones", "Activo", "Staff", "Seguidos", "Seguidores", "Likeados", "Likes", "Vistas"])

                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                usuaros_con_posts = 0
                pos = 2

                for usuario in usuarios:
                    color_active = green_fill if usuario.is_active else red_fill
                    color_staff = green_fill if usuario.is_staff else red_fill
                    usuaros_con_posts += 1  if Post.objects.filter(usuario=usuario).exists() else 0
                    len_posts = len(Post.objects.filter(usuario=usuario)) if usuaros_con_posts else 0
                    likieados = len(Like.objects.filter(usuario=usuario)) if Like.objects.filter(usuario=usuario).exists() else 0
                    likes = len(Like.objects.filter(del_usuario=usuario)) if Like.objects.filter(del_usuario=usuario).exists() else 0
                    vistas = len(Vistas.objects.filter(del_usuario=usuario)) if Vistas.objects.filter(del_usuario=usuario).exists() else 0
                    ws.append([
                        usuario.matricula,
                        usuario.nombre,
                        usuario.email,
                        usuario.fecha_de_entrada,
                        usuario.fecha_de_nacimiento,
                        len_posts,
                        usuario.is_active,
                        usuario.is_staff,
                        usuario.get_seguidores(),
                        usuario.get_seguidos(),
                        likieados,
                        likes,
                        vistas
                    ])

                    ws.cell(row=pos, column=7).fill = color_active
                    ws.cell(row=pos, column=8).fill = color_staff
                    pos += 1
                
                # Genera el gráfico de pastel
                pie = PieChart()
                #labels = Reference(ws, min_col=2, min_row=2, max_row=pos-1)
                #data = Reference(ws, min_col=6, min_row=2, max_row=pos-1)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.title = "Publicaciones por Usuario"
                pie.height = 10  # Ajusta la altura del gráfico
                pie.width = 15  # Ajusta el ancho del gráfico

                # Agrega el gráfico de pastel a la hoja de cálculo
                ws.add_chart(pie, "K1")

                # Guarda el archivo de Excel
                wb.save("sample.xlsx")

                return Response({"message": "Archivo Excel generado correctamente"}, status=status.HTTP_200_OK)
